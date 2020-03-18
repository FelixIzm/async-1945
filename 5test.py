import requests
import json
#from lxml.html import fromstring
import lxml.html as html
import hashlib
import os
from pathlib import Path
import tempfile
from openpyxl import Workbook

image_id=85942988
cookies = {}
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
cols = ['ID','Фамилия','Имя','Отчество','Дата рождения/Возраст','Место рождения','Дата и место призыва','Последнее место службы','Воинское звание','Судьба','Дата смерти','Первичное место захоронения']
path2save = os.path.join(BASE_DIR,str(image_id))
Path(path2save).mkdir(parents=True, exist_ok=True)



#####################################
def getStringHash(id):
    h = hashlib.md5(str(id).encode()+b'db76xdlrtxcxcghn7yusxjcdxsbtq1hnicnaspohh5tzbtgqjixzc5nmhybeh')
    p = h.hexdigest()
    return str(p)
#####################################
def parse_file (name_file):
    dict_ = {}
    f = open(name_file, 'r')
    s = f.read()
    dict_={}
    list_ = s.splitlines()
    for item in list_:
        items = item.split(":")
        dict_[items[0]] = items[1].lstrip()
    return dict_
#####################################
def make_str_cookie(cookies):
    str_cook = ''
    for key, value in cookies.items():
        str_cook += '{0}={1};'.format(key,value)
    return str_cook
#####################################

def get_lists(image_id):
    getimageinfo_url = 'https://obd-memorial.ru/html/getimageinfo?id={}'.format(image_id)
    info_url = 'https://obd-memorial.ru/html/info.htm?id={}'.format(image_id)
    list_id_images = []
    list_urls_infocards  = []
    res1 = requests.get(info_url)
    print(res1.status_code)
    if(res1.status_code==307):
        cookies = {}
        cookies['3fbe47cd30daea60fc16041479413da2']=res1.cookies['3fbe47cd30daea60fc16041479413da2']
        cookies['JSESSIONID']=res1.cookies['JSESSIONID']
        cookies['showimage']='0'
        img_info = 'https://obd-memorial.ru/html/getimageinfo?id={}'.format(image_id)
        response = requests.get(img_info)
        response_dict = json.loads(response.text)
        i=0
        for item in response_dict:
            i+=1
            #img_url="https://obd-memorial.ru/html/images3?id="+str(item['id'])+"&id1="+(getStringHash(item['id']))+"&path="+item['img']
            list_id_images.append({'id':item['id'],'img':item['img']})
            for id in item['mapData'].keys():
                info_url = 'https://obd-memorial.ru/html/info.htm?id='+str(id)
                list_urls_infocards.append(info_url)
    return(list_id_images,list_urls_infocards, cookies)

def get_info(info_url,cookies):
    global cols
    cookies['showimage']='0'
    #info_url = 'https://obd-memorial.ru/html/info.htm?id='+str(id)
    res3 = requests.get(info_url,cookies=cookies)
    doc = html.fromstring(res3.text)
    divs = {}
    for div in doc.find_class('card_parameter'):
        divs[div.getchildren()[0].text_content()] = div.getchildren()[1].text_content()
        #print ('%s: %s' % (div.getchildren()[0].text_content(), div.getchildren()[1].text_content()))
    list_col = []
    for col in cols:
        if(col in divs.keys()):
            list_col.append(divs[col])
        else:
            list_col.append('')
    #list_col[0] = id_scan
    list_col[0] = info_url[(info_url.index('=')+1):]
    return list_col

list_images, list_infocards, cookies = get_lists(image_id)


row_num = 1
workbook = Workbook()
# Get active worksheet/tab
worksheet = workbook.active
worksheet.title = 'Person'
columns = cols
for col_num, column_title in enumerate(columns, 1):
    cell = worksheet.cell(row=row_num, column=col_num)
    cell.value = column_title
#################################3
for url in list_infocards:
    row_num += 1
    row = get_info(url,cookies)
    #print('\t',id)
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
#################################3

workbook.save(filename =  path2save+"/"+str(image_id)+'_book.xlsx')
print('done')